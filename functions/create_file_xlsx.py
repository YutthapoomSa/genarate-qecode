import os
from openpyxl import Workbook

def create_file(data):
    # Specify the Excel file path
    config_folder = "./config/"
    excel_file_path = os.path.join(config_folder, "number.xlsx")

    # Create a new Workbook
    workbook = Workbook()

    # Select the first sheet
    sheet = workbook.active

    # Add a header to the single column
    headers = ["number"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Insert data into Excel
    for row, value in enumerate(data, start=2):  # Start from row 2 to leave space for headers
        sheet.cell(row=row, column=1, value=value)

    # Save the Excel file
    workbook.save(excel_file_path)

    return excel_file_path